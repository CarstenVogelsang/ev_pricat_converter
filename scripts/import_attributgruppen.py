#!/usr/bin/env python3
"""
Import-Script für NTG Attributgruppenschlüssel

Liest die 5-Ebenen Produktklassifikation aus der Excel-Datei
und importiert sie in die Attributgruppe-Tabelle.

Ausführung:
    uv run python scripts/import_attributgruppen.py

Struktur (1684+ Einträge):
    - Ebene 1: Hauptkategorie (17 Stück: Spielzeug, Schreibwaren, etc.)
    - Ebene 2: Unterkategorie (~100)
    - Ebene 3: Sub-Unterkategorie (~300)
    - Ebene 4: Produktgruppe (~800)
    - Ebene 5: Spezifischer Typ (1684 Blatt-Einträge)

NTG-Schlüssel Format: "101001001001000"
    = 1 (Default) + 01 (Ebene1) + 001 (Ebene2) + 001 (Ebene3) + 001 (Ebene4) + 000 (Ebene5)
"""
import os
import sys

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from app import create_app, db
from app.models import Attributgruppe


def clean_value(value):
    """Clean cell value - handle strings, numbers, None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        # Convert to string, pad codes appropriately
        return str(int(value))
    value = str(value).strip()
    return value if value else None


def pad_code(value, length):
    """Pad a code to the specified length with leading zeros."""
    if value is None:
        return None
    return str(value).zfill(length)


def main():
    """Main import function."""
    print("=" * 60)
    print("NTG Attributgruppenschlüssel Import")
    print("=" * 60)

    # Path to Excel file
    excel_path = 'docs/prd/module/PRD-009-produktdaten/NTG_Attributgruppenschluessel.xlsx'

    if not os.path.exists(excel_path):
        print(f"FEHLER: Excel-Datei nicht gefunden: {excel_path}")
        return 1

    # Create Flask app context
    app = create_app()

    with app.app_context():
        # Check if table already has data
        existing_count = Attributgruppe.query.count()
        if existing_count > 0:
            print(f"\n⚠️  Attributgruppe enthält bereits {existing_count} Einträge.")
            response = input("Alle löschen und neu importieren? (j/N): ")
            if response.lower() != 'j':
                print("Abbruch.")
                return 0

            # Delete all existing entries
            Attributgruppe.query.delete()
            db.session.commit()
            print(f"✓ {existing_count} Einträge gelöscht.\n")

        # Open Excel workbook
        print(f"Öffne: {excel_path}")
        wb = openpyxl.load_workbook(excel_path, read_only=True)

        # Get the main sheet
        sheet_name = 'NTG Attributgruppenschlüssel'
        if sheet_name not in wb.sheetnames:
            print(f"FEHLER: Sheet '{sheet_name}' nicht gefunden")
            print(f"Verfügbare Sheets: {wb.sheetnames}")
            return 1

        ws = wb[sheet_name]

        entries = []
        seen_keys = set()  # Track NTG-Schlüssel for duplicates
        skipped = 0

        print(f"Lese Sheet: {sheet_name}")

        # Skip header row (row 1), read data from row 2
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Column mapping:
            # 0: Default, 1: Ebene1 Code, 2: Ebene1 Name,
            # 3: Ebene2 Code, 4: Ebene2 Name,
            # 5: Ebene3 Code, 6: Ebene3 Name,
            # 7: Ebene4 Code, 8: Ebene4 Name,
            # 9: Ebene5 Code, 10: Ebene5 Name,
            # 11: NTG-Schlüssel

            ntg_schluessel = clean_value(row[11]) if len(row) > 11 else None

            # Skip if no NTG key
            if not ntg_schluessel:
                skipped += 1
                continue

            # Check for duplicates
            if ntg_schluessel in seen_keys:
                skipped += 1
                continue
            seen_keys.add(ntg_schluessel)

            # Extract and pad codes
            ebene_1_code = pad_code(clean_value(row[1]), 2)
            ebene_1_name = clean_value(row[2])
            ebene_2_code = pad_code(clean_value(row[3]), 3)
            ebene_2_name = clean_value(row[4])
            ebene_3_code = pad_code(clean_value(row[5]), 3)
            ebene_3_name = clean_value(row[6])
            ebene_4_code = pad_code(clean_value(row[7]), 3)
            ebene_4_name = clean_value(row[8])
            ebene_5_code = pad_code(clean_value(row[9]), 3)
            ebene_5_name = clean_value(row[10])

            entry = Attributgruppe(
                ntg_schluessel=ntg_schluessel,
                ebene_1_code=ebene_1_code,
                ebene_1_name=ebene_1_name,
                ebene_2_code=ebene_2_code,
                ebene_2_name=ebene_2_name,
                ebene_3_code=ebene_3_code,
                ebene_3_name=ebene_3_name,
                ebene_4_code=ebene_4_code,
                ebene_4_name=ebene_4_name,
                ebene_5_code=ebene_5_code,
                ebene_5_name=ebene_5_name,
                aktiv=True
            )
            entries.append(entry)

            # Progress indicator
            if len(entries) % 500 == 0:
                print(f"  ... {len(entries)} Einträge gelesen")

        print(f"  ✓ {len(entries)} Einträge gelesen")
        if skipped > 0:
            print(f"  ⚠️  {skipped} Zeilen übersprungen (leer/Duplikate)")

        # Add all entries to database
        print("\nSpeichere in Datenbank...")
        for entry in entries:
            db.session.add(entry)

        db.session.commit()

        print("-" * 60)
        print(f"✓ Import abgeschlossen: {len(entries)} Attributgruppen")

        # Summary by Ebene 1
        print("\n=== Zusammenfassung nach Hauptkategorie (Ebene 1) ===")
        hauptkategorien = Attributgruppe.get_hauptkategorien()
        for kat in hauptkategorien:
            print(f"  {kat['code']}: {kat['name']} ({kat['anzahl']} Einträge)")

    return 0


if __name__ == '__main__':
    sys.exit(main())
