"""XLSX Exporter Service.

Exports Lieferant, Hersteller, and Marke entities to Excel format.
"""
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.models import Lieferant, Hersteller, Marke


@dataclass
class XlsxExportResult:
    """Result of XLSX export."""
    success: bool
    output_path: Optional[Path] = None
    sheets_created: int = 0
    errors: list = None

    def __post_init__(self):
        if self.errors is None:
            self.errors = []


class XlsxExporter:
    """Exports entity data to XLSX format."""

    def __init__(self):
        self.header_font = Font(bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center')
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _style_header_row(self, ws, num_columns: int) -> None:
        """Apply styling to header row."""
        for col in range(1, num_columns + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

    def _auto_column_width(self, ws) -> None:
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    cell_length = len(str(cell.value)) if cell.value else 0
                    max_length = max(max_length, cell_length)
                except Exception:
                    pass

            # Set width with some padding, max 50 chars
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = max(adjusted_width, 10)

    def _add_lieferant_sheet(self, wb: Workbook, lieferant: Lieferant) -> None:
        """Add Lieferant sheet to workbook."""
        ws = wb.create_sheet('Lieferant')

        headers = ['GLN', 'VEDES-ID', 'Kurzbezeichnung', 'Aktiv', 'FTP Pfad Quelle',
                   'FTP Pfad Ziel', 'Elena Startdir', 'Elena Base URL', 'Letzte Konvertierung']
        ws.append(headers)
        self._style_header_row(ws, len(headers))

        # Add lieferant data
        ws.append([
            lieferant.gln,
            lieferant.vedes_id,
            lieferant.kurzbezeichnung,
            'Ja' if lieferant.aktiv else 'Nein',
            lieferant.ftp_pfad_quelle or '',
            lieferant.ftp_pfad_ziel or '',
            lieferant.elena_startdir or '',
            lieferant.elena_base_url or '',
            lieferant.letzte_konvertierung.strftime('%Y-%m-%d %H:%M') if lieferant.letzte_konvertierung else ''
        ])

        self._auto_column_width(ws)

    def _add_hersteller_sheet(self, wb: Workbook, hersteller_list: list[Hersteller]) -> None:
        """Add Hersteller sheet to workbook."""
        ws = wb.create_sheet('Hersteller')

        headers = ['ID', 'GLN', 'VEDES-ID', 'Kurzbezeichnung', 'Anzahl Marken']
        ws.append(headers)
        self._style_header_row(ws, len(headers))

        for h in hersteller_list:
            marken_count = h.marken.count() if hasattr(h, 'marken') else 0
            ws.append([
                h.id,
                h.gln,
                h.vedes_id or '',
                h.kurzbezeichnung,
                marken_count
            ])

        self._auto_column_width(ws)

    def _add_marken_sheet(self, wb: Workbook, marken_list: list[Marke]) -> None:
        """Add Marken sheet to workbook."""
        ws = wb.create_sheet('Marken')

        headers = ['ID', 'Kurzbezeichnung', 'GLN evendo', 'Hersteller-ID', 'Hersteller']
        ws.append(headers)
        self._style_header_row(ws, len(headers))

        for m in marken_list:
            hersteller_name = m.hersteller.kurzbezeichnung if m.hersteller else ''
            ws.append([
                m.id,
                m.kurzbezeichnung,
                m.gln_evendo,
                m.hersteller_id,
                hersteller_name
            ])

        self._auto_column_width(ws)

    def _add_summary_sheet(
        self,
        wb: Workbook,
        lieferant: Lieferant,
        hersteller_list: list[Hersteller],
        marken_list: list[Marke],
        article_count: int = 0
    ) -> None:
        """Add summary sheet to workbook."""
        ws = wb.active
        ws.title = 'Zusammenfassung'

        # Title
        ws['A1'] = 'PRICAT Konvertierung - Zusammenfassung'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:C1')

        # Export info
        ws['A3'] = 'Export-Datum:'
        ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        ws['A4'] = 'Lieferant:'
        ws['B4'] = lieferant.kurzbezeichnung if lieferant else '-'

        ws['A5'] = 'GLN:'
        ws['B5'] = lieferant.gln if lieferant else '-'

        # Statistics
        ws['A7'] = 'Statistik'
        ws['A7'].font = Font(bold=True)

        ws['A8'] = 'Anzahl Artikel:'
        ws['B8'] = article_count

        ws['A9'] = 'Anzahl Hersteller:'
        ws['B9'] = len(hersteller_list)

        ws['A10'] = 'Anzahl Marken:'
        ws['B10'] = len(marken_list)

        # Style
        for row in range(3, 11):
            ws[f'A{row}'].font = Font(bold=True)

        self._auto_column_width(ws)

    def export_entities(
        self,
        lieferant: Lieferant,
        hersteller: list[Hersteller],
        marken: list[Marke],
        output_path: Path,
        article_count: int = 0
    ) -> XlsxExportResult:
        """
        Export entities to XLSX file.

        Args:
            lieferant: Supplier entity
            hersteller: List of manufacturer entities
            marken: List of brand entities
            output_path: Path for output XLSX file
            article_count: Optional article count for summary

        Returns:
            XlsxExportResult with success status
        """
        result = XlsxExportResult(success=False)

        try:
            # Ensure output directory exists
            output_path.parent.mkdir(parents=True, exist_ok=True)

            wb = Workbook()

            # Add sheets
            self._add_summary_sheet(wb, lieferant, hersteller, marken, article_count)
            result.sheets_created += 1

            if lieferant:
                self._add_lieferant_sheet(wb, lieferant)
                result.sheets_created += 1

            if hersteller:
                self._add_hersteller_sheet(wb, hersteller)
                result.sheets_created += 1

            if marken:
                self._add_marken_sheet(wb, marken)
                result.sheets_created += 1

            # Save
            wb.save(output_path)

            result.success = True
            result.output_path = output_path

        except Exception as e:
            result.errors.append(f"XLSX export failed: {str(e)}")

        return result


def generate_xlsx_filename(lieferant_vedes_id: str) -> str:
    """
    Generate XLSX export filename.

    Args:
        lieferant_vedes_id: VEDES ID of supplier

    Returns:
        Filename like 'entities_0000001872_20250103_143052.xlsx'
    """
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return f"entities_{lieferant_vedes_id}_{timestamp}.xlsx"
