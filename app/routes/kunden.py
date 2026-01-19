"""Kunden (Customer) routes for Lead&Kundenreport app."""
from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify
from flask_login import login_required, current_user
from flask_wtf import FlaskForm
from wtforms import StringField, TextAreaField, BooleanField, SelectField
from wtforms.validators import DataRequired, URL, Optional

from sqlalchemy import func

from app import db
from app.models import Kunde, KundeCI, KundeApiNutzung, User, Branche, Verband, KundeBranche, KundeVerband
from app.models import BranchenRolle, KundeBranchenRolle, PasswordToken, KundeBenutzer, KundeTyp
from app.services import FirecrawlService, get_password_service
from app.services.logging_service import log_mittel
from app.routes.auth import mitarbeiter_required

kunden_bp = Blueprint('kunden', __name__, url_prefix='/kunden')


class KundeForm(FlaskForm):
    """Form for creating/editing Kunde."""
    firmierung = StringField('Firmierung', validators=[DataRequired()])
    ev_kdnr = StringField('e-vendo Kundennummer', validators=[Optional()])
    strasse = StringField('Straße', validators=[Optional()])
    plz = StringField('PLZ', validators=[Optional()])
    ort = StringField('Ort', validators=[Optional()])
    land = StringField('Land', validators=[Optional()])
    adresse = TextAreaField('Adresse (alt)', validators=[Optional()])
    website_url = StringField('Website URL', validators=[Optional(), URL()])
    shop_url = StringField('Online-Shop URL', validators=[Optional(), URL()])
    telefon = StringField('Telefon', validators=[Optional()])
    email = StringField('E-Mail', validators=[Optional()])

    # Kommunikationsstil (Firmen-Standard, Anrede ist jetzt am User)
    kommunikation_stil = SelectField('Kommunikationsstil (Standard)', choices=[
        ('foermlich', 'Förmlich (Sie)'),
        ('locker', 'Locker (Du)'),
    ], default='foermlich')

    notizen = TextAreaField('Notizen', validators=[Optional()])
    aktiv = BooleanField('Aktiv', default=True)

    # Typ: Kunde oder Lead (nur bei Neuanlage sichtbar)
    typ = SelectField('Typ', choices=[
        ('kunde', 'Kunde'),
        ('lead', 'Lead'),
    ], default='kunde')


@kunden_bp.route('/')
@login_required
@mitarbeiter_required
def liste():
    """List all Kunden/Leads with optional filters."""
    # Get filter parameters
    branche_id = request.args.get('branche', type=int)
    verband_id = request.args.get('verband', type=int)
    status = request.args.get('status', 'alle')
    typ_filter = request.args.get('typ', 'alle')  # kunde, lead, alle

    # Base query - always exclude test customers
    query = Kunde.query.filter(Kunde.typ != KundeTyp.TESTKUNDE.value)

    # Apply filters
    if branche_id:
        query = query.join(KundeBranche).filter(KundeBranche.branche_id == branche_id)
    if verband_id:
        query = query.join(KundeVerband).filter(KundeVerband.verband_id == verband_id)
    if status == 'aktiv':
        query = query.filter(Kunde.aktiv == True)
    elif status == 'inaktiv':
        query = query.filter(Kunde.aktiv == False)

    # Typ-Filter (Kunde vs Lead)
    if typ_filter == 'kunde':
        query = query.filter(Kunde.typ == KundeTyp.KUNDE.value)
    elif typ_filter == 'lead':
        query = query.filter(Kunde.typ == KundeTyp.LEAD.value)

    kunden = query.order_by(Kunde.firmierung).all()

    # Get all Branchen and Verbände for filter dropdowns
    alle_branchen = Branche.query.filter_by(aktiv=True).order_by(Branche.sortierung).all()
    alle_verbaende = Verband.query.filter_by(aktiv=True).order_by(Verband.name).all()

    return render_template(
        'kunden/liste.html',
        kunden=kunden,
        alle_branchen=alle_branchen,
        alle_verbaende=alle_verbaende,
        filter_branche=branche_id,
        filter_verband=verband_id,
        filter_status=status,
        filter_typ=typ_filter
    )


@kunden_bp.route('/neu', methods=['GET', 'POST'])
@login_required
@mitarbeiter_required
def neu():
    """Create new Kunde."""
    form = KundeForm()

    if form.validate_on_submit():
        kunde = Kunde(
            firmierung=form.firmierung.data,
            ev_kdnr=form.ev_kdnr.data or None,
            strasse=form.strasse.data,
            plz=form.plz.data,
            ort=form.ort.data,
            land=form.land.data or 'Deutschland',
            adresse=form.adresse.data,
            telefon=form.telefon.data or None,
            email=form.email.data or None,
            website_url=form.website_url.data or None,
            shop_url=form.shop_url.data or None,
            kommunikation_stil=form.kommunikation_stil.data,
            notizen=form.notizen.data,
            aktiv=form.aktiv.data,
            typ=form.typ.data  # Kunde oder Lead
        )
        db.session.add(kunde)
        db.session.commit()

        # Passende Meldung je nach Typ
        typ_label = 'Lead' if kunde.is_lead else 'Kunde'
        flash(f'{typ_label} "{kunde.firmierung}" wurde angelegt.', 'success')
        return redirect(url_for('kunden.detail', id=kunde.id))

    return render_template('kunden/form.html', form=form, titel='Neuer Kunde', kunde=None)


@kunden_bp.route('/<int:id>')
@login_required
@mitarbeiter_required
def detail(id):
    """View Kunde detail."""
    kunde = Kunde.query.get_or_404(id)

    # Hauptbranchen für Hauptbranche-Dropdown
    hauptbranchen = Branche.query.filter(
        Branche.aktiv == True,
        Branche.parent_id == None
    ).order_by(Branche.sortierung).all()

    # Branchenmodell V2: Nur Unterbranchen der Hauptbranche anzeigen
    # Erst wenn Hauptbranche gesetzt ist, können Unterbranchen zugeordnet werden
    if kunde.hauptbranche_id:
        alle_unterbranchen = Branche.query.filter(
            Branche.aktiv == True,
            Branche.parent_id == kunde.hauptbranche_id
        ).order_by(Branche.sortierung).all()
    else:
        alle_unterbranchen = []  # Keine Unterbranchen ohne Hauptbranche

    alle_verbaende = Verband.query.filter_by(aktiv=True).order_by(Verband.name).all()

    # Get IDs of assigned Branchen and Verbände for the template
    kunde_branchen_ids = {kb.branche_id for kb in kunde.branchen}
    kunde_primaer_ids = {kb.branche_id for kb in kunde.branchen if kb.ist_primaer}
    kunde_verbaende_ids = {kv.verband_id for kv in kunde.verbaende}

    # Branchenmodell V2: Rollen pro Branche für den Kunden
    # Dict: branche_id -> set(branchenrolle_id)
    kunde_rollen_ids = {}
    for kbr in kunde.branchenrollen:
        if kbr.branche_id not in kunde_rollen_ids:
            kunde_rollen_ids[kbr.branche_id] = set()
        kunde_rollen_ids[kbr.branche_id].add(kbr.branchenrolle_id)

    return render_template(
        'kunden/detail.html',
        kunde=kunde,
        alle_branchen=alle_unterbranchen,  # Backward-compatible name
        hauptbranchen=hauptbranchen,
        alle_verbaende=alle_verbaende,
        kunde_branchen_ids=kunde_branchen_ids,
        kunde_primaer_ids=kunde_primaer_ids,
        kunde_verbaende_ids=kunde_verbaende_ids,
        kunde_rollen_ids=kunde_rollen_ids
    )


@kunden_bp.route('/<int:id>/bearbeiten', methods=['GET', 'POST'])
@login_required
@mitarbeiter_required
def bearbeiten(id):
    """Edit Kunde."""
    kunde = Kunde.query.get_or_404(id)
    form = KundeForm(obj=kunde)

    if form.validate_on_submit():
        kunde.firmierung = form.firmierung.data
        kunde.ev_kdnr = form.ev_kdnr.data or None
        kunde.strasse = form.strasse.data
        kunde.plz = form.plz.data
        kunde.ort = form.ort.data
        kunde.land = form.land.data or 'Deutschland'
        kunde.adresse = form.adresse.data
        kunde.telefon = form.telefon.data or None
        kunde.email = form.email.data or None
        kunde.website_url = form.website_url.data or None
        kunde.shop_url = form.shop_url.data or None
        kunde.kommunikation_stil = form.kommunikation_stil.data
        kunde.notizen = form.notizen.data
        kunde.aktiv = form.aktiv.data
        db.session.commit()
        flash(f'Kunde "{kunde.firmierung}" wurde aktualisiert.', 'success')
        return redirect(url_for('kunden.detail', id=kunde.id))

    return render_template('kunden/form.html', form=form, titel='Kunde bearbeiten', kunde=kunde)


@kunden_bp.route('/<int:id>/loeschen', methods=['POST'])
@login_required
@mitarbeiter_required
def loeschen(id):
    """Delete Kunde."""
    kunde = Kunde.query.get_or_404(id)
    firmierung = kunde.firmierung
    db.session.delete(kunde)
    db.session.commit()
    flash(f'Kunde "{firmierung}" wurde geloescht.', 'success')
    return redirect(url_for('kunden.liste'))


@kunden_bp.route('/<int:id>/analyse', methods=['POST'])
@login_required
@mitarbeiter_required
def analyse(id):
    """Trigger Firecrawl website analysis."""
    kunde = Kunde.query.get_or_404(id)

    if not kunde.website_url:
        flash('Website-URL muss gesetzt sein für die Analyse.', 'warning')
        return redirect(url_for('kunden.detail', id=id))

    firecrawl_service = FirecrawlService()
    result = firecrawl_service.analyze_branding(kunde, user_id=current_user.id)

    if result.success:
        flash('Website-Analyse erfolgreich abgeschlossen.', 'success')
    else:
        flash(f'Analyse fehlgeschlagen: {result.error}', 'danger')

    return redirect(url_for('kunden.detail', id=id))


@kunden_bp.route('/<int:id>/reparse-logo', methods=['POST'])
@login_required
@mitarbeiter_required
def reparse_logo(id):
    """Re-extract logo from stored raw_response without new API call."""
    kunde = Kunde.query.get_or_404(id)

    if not kunde.ci or not kunde.ci.raw_response:
        flash('Kein Raw Response vorhanden. Bitte erst Website-Analyse durchfuehren.', 'warning')
        return redirect(url_for('kunden.detail', id=id))

    logo_url = FirecrawlService.reparse_logo_from_raw(kunde.ci)

    if logo_url:
        flash(f'Logo erfolgreich extrahiert.', 'success')
    else:
        flash('Kein Logo im Raw Response gefunden.', 'warning')

    return redirect(url_for('kunden.detail', id=id))


@kunden_bp.route('/<int:id>/projekt')
@login_required
@mitarbeiter_required
def projekt(id):
    """View Kunde project page with activities and API costs."""
    kunde = Kunde.query.get_or_404(id)

    # Get API usage for this Kunde (all users)
    api_nutzungen = KundeApiNutzung.query.filter_by(kunde_id=id)\
        .order_by(KundeApiNutzung.created_at.desc()).all()

    # Calculate totals
    totals = db.session.query(
        func.sum(KundeApiNutzung.credits_used).label('total_credits'),
        func.sum(KundeApiNutzung.kosten_euro).label('total_kosten'),
        func.count(KundeApiNutzung.id).label('anzahl_calls')
    ).filter(KundeApiNutzung.kunde_id == id).first()

    # Usage per user
    per_user = db.session.query(
        User.id,
        User.vorname,
        User.nachname,
        func.sum(KundeApiNutzung.credits_used).label('credits'),
        func.sum(KundeApiNutzung.kosten_euro).label('kosten'),
        func.count(KundeApiNutzung.id).label('calls')
    ).join(User, KundeApiNutzung.user_id == User.id)\
        .filter(KundeApiNutzung.kunde_id == id)\
        .group_by(User.id, User.vorname, User.nachname).all()

    return render_template(
        'kunden/projekt.html',
        kunde=kunde,
        api_nutzungen=api_nutzungen,
        totals=totals,
        per_user=per_user
    )


# ===== Hauptbranche Endpoint =====

@kunden_bp.route('/<int:id>/hauptbranche', methods=['POST'])
@login_required
@mitarbeiter_required
def set_hauptbranche(id):
    """Set the Hauptbranche for a Kunde.

    Expects JSON: { "hauptbranche_id": 123 } or { "hauptbranche_id": null }
    """
    kunde = Kunde.query.get_or_404(id)
    data = request.get_json()

    if data is None:
        return jsonify({'success': False, 'error': 'JSON erforderlich'}), 400

    hauptbranche_id = data.get('hauptbranche_id')

    if hauptbranche_id is None:
        # Hauptbranche entfernen
        kunde.hauptbranche_id = None
        db.session.commit()
        return jsonify({
            'success': True,
            'message': 'Hauptbranche entfernt',
            'hauptbranche_id': None
        })

    # Validieren: Muss Hauptbranche sein (parent_id=NULL)
    branche = Branche.query.get(hauptbranche_id)
    if not branche:
        return jsonify({'success': False, 'error': 'Branche nicht gefunden'}), 404

    if branche.parent_id is not None:
        return jsonify({
            'success': False,
            'error': 'Nur Hauptbranchen (HANDEL, HANDWERK, etc.) erlaubt'
        }), 400

    # Bei Hauptbranche-Wechsel: Bestehende Unterbranche-Zuordnungen bereinigen?
    # Für jetzt behalten wir sie - User kann manuell bereinigen
    alte_hauptbranche_id = kunde.hauptbranche_id

    kunde.hauptbranche_id = hauptbranche_id
    db.session.commit()

    return jsonify({
        'success': True,
        'message': f'Hauptbranche auf "{branche.name}" gesetzt',
        'hauptbranche_id': hauptbranche_id,
        'hauptbranche_name': branche.name
    })


@kunden_bp.route('/<int:id>/hauptbranche', methods=['DELETE'])
@login_required
@mitarbeiter_required
def delete_hauptbranche(id):
    """Delete the Hauptbranche and cascade-delete all Unterbranche-Zuordnungen.

    This is a destructive action that removes:
    - The Hauptbranche assignment
    - All KundeBranche entries for this Kunde
    - All KundeBranchenRolle entries for this Kunde
    """
    kunde = Kunde.query.get_or_404(id)

    if not kunde.hauptbranche_id:
        return jsonify({
            'success': False,
            'error': 'Keine Hauptbranche vorhanden'
        }), 400

    alte_hauptbranche_name = kunde.hauptbranche.name if kunde.hauptbranche else "Unbekannt"

    # Count affected entries for response
    deleted_branches_count = KundeBranche.query.filter_by(kunde_id=id).count()
    deleted_roles_count = KundeBranchenRolle.query.filter_by(kunde_id=id).count()

    # Delete all KundeBranchenRolle entries for this Kunde
    KundeBranchenRolle.query.filter_by(kunde_id=id).delete()

    # Delete all KundeBranche entries for this Kunde
    KundeBranche.query.filter_by(kunde_id=id).delete()

    # Remove Hauptbranche
    kunde.hauptbranche_id = None

    # Log the event
    log_mittel(
        modul='kunden',
        aktion='hauptbranche_geloescht',
        details=f'Hauptbranche "{alte_hauptbranche_name}" und {deleted_branches_count} Unterbranchen-Zuordnungen entfernt für Kunde "{kunde.firmierung}"',
        entity_type='Kunde',
        entity_id=id
    )

    db.session.commit()

    return jsonify({
        'success': True,
        'message': f'Hauptbranche "{alte_hauptbranche_name}" und {deleted_branches_count} Unterbranchen-Zuordnungen entfernt',
        'deleted_branches': deleted_branches_count,
        'deleted_roles': deleted_roles_count
    })


# ===== AJAX Endpoints for Branchen/Verbände =====

@kunden_bp.route('/<int:id>/branche/<int:branche_id>', methods=['POST'])
@login_required
@mitarbeiter_required
def toggle_branche(id, branche_id):
    """Toggle Branche assignment for a Kunde."""
    kunde = Kunde.query.get_or_404(id)
    branche = Branche.query.get_or_404(branche_id)

    existing = KundeBranche.query.filter_by(
        kunde_id=id, branche_id=branche_id
    ).first()

    if existing:
        # Remove assignment
        db.session.delete(existing)
        db.session.commit()
        return jsonify({'success': True, 'action': 'removed'})
    else:
        # Add assignment
        kb = KundeBranche(kunde_id=id, branche_id=branche_id, ist_primaer=False)
        db.session.add(kb)
        db.session.commit()
        return jsonify({'success': True, 'action': 'added'})


@kunden_bp.route('/<int:id>/branche/<int:branche_id>/primaer', methods=['POST'])
@login_required
@mitarbeiter_required
def toggle_primaer_branche(id, branche_id):
    """Toggle primary status for a Branche assignment. Max 3 primary branches."""
    kunde = Kunde.query.get_or_404(id)

    existing = KundeBranche.query.filter_by(
        kunde_id=id, branche_id=branche_id
    ).first()

    if not existing:
        # Branch not assigned yet - assign and make primary
        primaer_count = KundeBranche.query.filter_by(kunde_id=id, ist_primaer=True).count()
        if primaer_count >= 3:
            return jsonify({
                'success': False,
                'error': 'Maximal 3 Primärbranchen erlaubt'
            })
        kb = KundeBranche(kunde_id=id, branche_id=branche_id, ist_primaer=True)
        db.session.add(kb)
        db.session.commit()
        return jsonify({'success': True, 'action': 'added_primary'})
    else:
        if existing.ist_primaer:
            # Remove primary status
            existing.ist_primaer = False
            db.session.commit()
            return jsonify({'success': True, 'action': 'removed_primary'})
        else:
            # Make primary (check limit)
            primaer_count = KundeBranche.query.filter_by(kunde_id=id, ist_primaer=True).count()
            if primaer_count >= 3:
                return jsonify({
                    'success': False,
                    'error': 'Maximal 3 Primärbranchen erlaubt'
                })
            existing.ist_primaer = True
            db.session.commit()
            return jsonify({'success': True, 'action': 'set_primary'})


@kunden_bp.route('/<int:id>/verband/<int:verband_id>', methods=['POST'])
@login_required
@mitarbeiter_required
def toggle_verband(id, verband_id):
    """Toggle Verband membership for a Kunde."""
    kunde = Kunde.query.get_or_404(id)
    verband = Verband.query.get_or_404(verband_id)

    existing = KundeVerband.query.filter_by(
        kunde_id=id, verband_id=verband_id
    ).first()

    if existing:
        # Remove membership
        db.session.delete(existing)
        db.session.commit()
        return jsonify({'success': True, 'action': 'removed'})
    else:
        # Add membership
        kv = KundeVerband(kunde_id=id, verband_id=verband_id)
        db.session.add(kv)
        db.session.commit()
        return jsonify({'success': True, 'action': 'added'})


# ===== Branchenmodell V2: Rollen-Endpunkte =====

@kunden_bp.route('/<int:id>/branche/<int:branche_id>/rollen', methods=['GET'])
@login_required
@mitarbeiter_required
def branche_rollen_get(id, branche_id):
    """Get available and assigned roles for a Kunde-Branche combination."""
    kunde = Kunde.query.get_or_404(id)
    branche = Branche.query.get_or_404(branche_id)

    # Zulässige Rollen für diese Branche
    zulaessige_rollen = branche.zulaessige_branchenrollen

    # Bereits zugewiesene Rollen für diesen Kunden in dieser Branche
    zugewiesene_ids = {
        kbr.branchenrolle_id
        for kbr in kunde.branchenrollen
        if kbr.branche_id == branche_id
    }

    return jsonify({
        'success': True,
        'branche': {
            'id': branche.id,
            'name': branche.name,
            'icon': branche.icon
        },
        'rollen': [
            {
                'id': rolle.id,
                'code': rolle.code,
                'name': rolle.name,
                'icon': rolle.icon,
                'beschreibung': rolle.beschreibung,
                'zugewiesen': rolle.id in zugewiesene_ids
            }
            for rolle in zulaessige_rollen
        ]
    })


@kunden_bp.route('/<int:id>/branche/<int:branche_id>/rollen', methods=['POST'])
@login_required
@mitarbeiter_required
def branche_rollen_set(id, branche_id):
    """Set roles for a Kunde-Branche combination.

    Expects JSON: { "rollen_ids": [1, 2, 3] }
    """
    kunde = Kunde.query.get_or_404(id)
    branche = Branche.query.get_or_404(branche_id)

    data = request.get_json()
    if not data or 'rollen_ids' not in data:
        return jsonify({'success': False, 'error': 'rollen_ids erforderlich'}), 400

    rollen_ids = set(data['rollen_ids'])

    # Zulässige Rollen-IDs für diese Branche
    zulaessige_ids = {r.id for r in branche.zulaessige_branchenrollen}

    # Validierung: Nur zulässige Rollen akzeptieren
    ungueltige = rollen_ids - zulaessige_ids
    if ungueltige:
        return jsonify({
            'success': False,
            'error': f'Ungültige Rollen-IDs: {list(ungueltige)}'
        }), 400

    # Bestehende Zuordnungen für diesen Kunden+Branche löschen
    KundeBranchenRolle.query.filter_by(
        kunde_id=id,
        branche_id=branche_id
    ).delete()

    # Neue Zuordnungen anlegen
    for rolle_id in rollen_ids:
        kbr = KundeBranchenRolle(
            kunde_id=id,
            branche_id=branche_id,
            branchenrolle_id=rolle_id
        )
        db.session.add(kbr)

    db.session.commit()

    return jsonify({
        'success': True,
        'message': f'{len(rollen_ids)} Rolle(n) zugewiesen',
        'rollen_count': len(rollen_ids)
    })


# ===== PRD-006: User-Erstellung für Kunden =====

@kunden_bp.route('/<int:id>/user/create', methods=['POST'])
@login_required
@mitarbeiter_required
def create_user(id):
    """Create a user account for a Kunde.

    Expects JSON:
    {
        "email": "kunde@example.com",
        "vorname": "Max",
        "nachname": "Mustermann"
    }

    The first user created for a Kunde automatically becomes Hauptbenutzer.
    Returns JSON with user info and password token.
    """
    kunde = Kunde.query.get_or_404(id)

    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'JSON erforderlich'}), 400

    email = data.get('email', '').strip()
    vorname = data.get('vorname', '').strip()
    nachname = data.get('nachname', '').strip()

    if not email or not vorname or not nachname:
        return jsonify({
            'success': False,
            'error': 'E-Mail, Vorname und Nachname erforderlich'
        }), 400

    # Check if email is already in use
    if User.query.filter_by(email=email.lower()).first():
        return jsonify({
            'success': False,
            'error': 'Diese E-Mail-Adresse wird bereits verwendet'
        }), 400

    service = get_password_service()
    result = service.create_user_for_kunde(kunde, email, vorname, nachname)

    if not result.success:
        return jsonify({
            'success': False,
            'error': result.error
        }), 400

    # Log the event
    is_hauptbenutzer = result.is_hauptbenutzer if hasattr(result, 'is_hauptbenutzer') else True
    log_mittel(
        modul='kunden',
        aktion='user_erstellt',
        details=f'User-Account für Kunde "{kunde.firmierung}" erstellt: {email}' +
                (' (Hauptbenutzer)' if is_hauptbenutzer else ''),
        entity_type='Kunde',
        entity_id=id
    )

    return jsonify({
        'success': True,
        'user': result.user.to_dict(),
        'token_id': result.password_token.id,
        'is_hauptbenutzer': is_hauptbenutzer,
        'message': f'User-Account für {email} erstellt'
    })


@kunden_bp.route('/<int:id>/user/send-credentials', methods=['POST'])
@login_required
@mitarbeiter_required
def send_credentials(id):
    """Send credential emails to the Kunde's user.

    Sends two separate emails:
    1. Portal-URL and username
    2. Password reveal link

    The password token must still be valid (not revealed, not expired).
    """
    kunde = Kunde.query.get_or_404(id)

    if not kunde.user_id:
        return jsonify({
            'success': False,
            'error': 'Kunde hat keinen User-Account'
        }), 400

    user = kunde.user

    # Find a valid password token
    password_token = PasswordToken.query.filter_by(
        user_id=user.id
    ).order_by(PasswordToken.created_at.desc()).first()

    if not password_token:
        return jsonify({
            'success': False,
            'error': 'Kein Passwort-Token vorhanden. Bitte neuen User erstellen.'
        }), 400

    if not password_token.is_valid:
        if password_token.is_revealed:
            error = 'Passwort wurde bereits angezeigt. Bitte neuen Token erstellen.'
        else:
            error = 'Passwort-Token ist abgelaufen. Bitte neuen Token erstellen.'
        return jsonify({
            'success': False,
            'error': error
        }), 400

    service = get_password_service()
    result = service.send_credentials(user, password_token)

    if not result.success:
        return jsonify({
            'success': False,
            'error': result.error
        }), 400

    # Log the event
    log_mittel(
        modul='kunden',
        aktion='zugangsdaten_gesendet',
        details=f'Zugangsdaten an {user.email} gesendet für Kunde "{kunde.firmierung}"',
        entity_type='Kunde',
        entity_id=id
    )

    return jsonify({
        'success': True,
        'message': f'Zugangsdaten an {user.email} gesendet'
    })


@kunden_bp.route('/<int:id>/user/new-token', methods=['POST'])
@login_required
@mitarbeiter_required
def create_new_token(id):
    """Create a new password token for the Kunde's user.

    This generates a new random password and token for re-sending credentials.
    """
    kunde = Kunde.query.get_or_404(id)

    if not kunde.user_id:
        return jsonify({
            'success': False,
            'error': 'Kunde hat keinen User-Account'
        }), 400

    user = kunde.user
    service = get_password_service()

    # Generate new password
    new_password = service.generate_secure_password()
    user.set_password(new_password)

    # Create new token
    password_token = PasswordToken.create_for_user(
        user_id=user.id,
        password_plain=new_password
    )
    db.session.add(password_token)
    db.session.commit()

    # Log the event
    log_mittel(
        modul='kunden',
        aktion='neues_passwort',
        details=f'Neues Passwort für User {user.email} erstellt',
        entity_type='Kunde',
        entity_id=id
    )

    return jsonify({
        'success': True,
        'token_id': password_token.id,
        'message': 'Neues Passwort erstellt. Zugangsdaten können jetzt gesendet werden.'
    })


@kunden_bp.route('/<int:id>/available-users')
@login_required
@mitarbeiter_required
def available_users(id):
    """Get list of available users for assignment to a Kunde.

    Returns users not already assigned to THIS Kunde, sorted by domain match.
    If the Kunde has an email, users with matching domain are prioritized.
    """
    kunde = Kunde.query.get_or_404(id)

    # Get IDs of users already assigned to THIS Kunde
    already_assigned_ids = {kb.user_id for kb in kunde.benutzer_zuordnungen}

    # Find active users who are not already assigned to this Kunde
    available = User.query.filter(
        User.aktiv == True,
        ~User.id.in_(already_assigned_ids) if already_assigned_ids else True
    ).all()

    # Extract domain from Kunde email if available
    matching_domain = None
    if kunde.email and '@' in kunde.email:
        matching_domain = kunde.email.split('@')[1].lower()

    # Build user list with domain_match flag
    user_list = []
    for user in available:
        user_domain = user.email.split('@')[1].lower() if '@' in user.email else ''
        domain_match = matching_domain and user_domain == matching_domain

        # Check if user is assigned to other Kunden
        other_kunden = [zuo.kunde.firmierung for zuo in user.kunde_zuordnungen
                        if zuo.kunde_id != id]

        user_list.append({
            'id': user.id,
            'email': user.email,
            'full_name': user.full_name,
            'rolle': user.rolle,
            'domain_match': domain_match,
            'andere_kunden': other_kunden
        })

    # Sort: domain matches first, then by name
    user_list.sort(key=lambda u: (not u['domain_match'], u['full_name'].lower()))

    return jsonify({
        'success': True,
        'users': user_list,
        'matching_domain': matching_domain,
        'total': len(user_list)
    })


@kunden_bp.route('/<int:id>/user/assign', methods=['POST'])
@login_required
@mitarbeiter_required
def assign_user(id):
    """Assign an existing user to a Kunde.

    Expects JSON:
    {
        "user_id": 123
    }

    The first user assigned becomes Hauptbenutzer automatically.
    """
    kunde = Kunde.query.get_or_404(id)

    data = request.get_json()
    if not data or 'user_id' not in data:
        return jsonify({'success': False, 'error': 'user_id erforderlich'}), 400

    user_id = data['user_id']
    user = User.query.get(user_id)

    if not user:
        return jsonify({'success': False, 'error': 'Benutzer nicht gefunden'}), 404

    # Check if user is already assigned to THIS Kunde
    existing = KundeBenutzer.query.filter_by(kunde_id=id, user_id=user_id).first()
    if existing:
        return jsonify({
            'success': False,
            'error': f'Benutzer ist diesem Kunden bereits zugewiesen'
        }), 400

    # First user becomes Hauptbenutzer
    is_first_user = len(kunde.benutzer_zuordnungen) == 0
    is_hauptbenutzer = is_first_user

    # Create KundeBenutzer entry
    kunde_benutzer = KundeBenutzer(
        kunde_id=id,
        user_id=user_id,
        ist_hauptbenutzer=is_hauptbenutzer
    )
    db.session.add(kunde_benutzer)

    # Also update legacy user_id for backward compatibility (if first user)
    if is_first_user:
        kunde.user_id = user_id

    db.session.commit()

    # Log the event
    log_mittel(
        modul='kunden',
        aktion='user_zugewiesen',
        details=f'Benutzer {user.full_name} ({user.email}) wurde Kunde "{kunde.firmierung}" zugewiesen' +
                (' (Hauptbenutzer)' if is_hauptbenutzer else ''),
        entity_type='Kunde',
        entity_id=id
    )

    return jsonify({
        'success': True,
        'message': f'Benutzer {user.full_name} wurde zugewiesen',
        'is_hauptbenutzer': is_hauptbenutzer,
        'user': {
            'id': user.id,
            'email': user.email,
            'full_name': user.full_name
        }
    })


@kunden_bp.route('/<int:id>/user/<int:user_id>/remove', methods=['POST'])
@login_required
@mitarbeiter_required
def remove_user(id, user_id):
    """Remove a user assignment from a Kunde.

    If the removed user was Hauptbenutzer, the next user becomes Hauptbenutzer.
    """
    kunde = Kunde.query.get_or_404(id)
    user = User.query.get_or_404(user_id)

    # Find the assignment
    zuordnung = KundeBenutzer.query.filter_by(kunde_id=id, user_id=user_id).first()
    if not zuordnung:
        return jsonify({
            'success': False,
            'error': 'Benutzer ist diesem Kunden nicht zugewiesen'
        }), 404

    was_hauptbenutzer = zuordnung.ist_hauptbenutzer

    # Remove the assignment
    db.session.delete(zuordnung)

    # If was Hauptbenutzer, promote the next user
    if was_hauptbenutzer:
        remaining = KundeBenutzer.query.filter_by(kunde_id=id).first()
        if remaining:
            remaining.ist_hauptbenutzer = True
            kunde.user_id = remaining.user_id  # Update legacy field
        else:
            kunde.user_id = None  # No users left

    db.session.commit()

    # Log the event
    log_mittel(
        modul='kunden',
        aktion='user_entfernt',
        details=f'Benutzer {user.full_name} ({user.email}) wurde von Kunde "{kunde.firmierung}" entfernt',
        entity_type='Kunde',
        entity_id=id
    )

    return jsonify({
        'success': True,
        'message': f'Benutzer {user.full_name} wurde entfernt',
        'new_hauptbenutzer_id': kunde.hauptbenutzer.id if kunde.hauptbenutzer else None
    })


@kunden_bp.route('/<int:id>/user/<int:user_id>/set-hauptbenutzer', methods=['POST'])
@login_required
@mitarbeiter_required
def set_hauptbenutzer(id, user_id):
    """Set a user as Hauptbenutzer for a Kunde.

    Only one user can be Hauptbenutzer - the previous one is demoted.
    """
    kunde = Kunde.query.get_or_404(id)
    user = User.query.get_or_404(user_id)

    # Find the assignment
    zuordnung = KundeBenutzer.query.filter_by(kunde_id=id, user_id=user_id).first()
    if not zuordnung:
        return jsonify({
            'success': False,
            'error': 'Benutzer ist diesem Kunden nicht zugewiesen'
        }), 404

    if zuordnung.ist_hauptbenutzer:
        return jsonify({
            'success': False,
            'error': 'Benutzer ist bereits Hauptbenutzer'
        }), 400

    # Remove Hauptbenutzer flag from current
    current_hauptbenutzer = KundeBenutzer.query.filter_by(
        kunde_id=id, ist_hauptbenutzer=True
    ).first()
    if current_hauptbenutzer:
        current_hauptbenutzer.ist_hauptbenutzer = False

    # Set new Hauptbenutzer
    zuordnung.ist_hauptbenutzer = True

    # Update legacy user_id field
    kunde.user_id = user_id

    db.session.commit()

    # Log the event
    log_mittel(
        modul='kunden',
        aktion='hauptbenutzer_geaendert',
        details=f'Benutzer {user.full_name} wurde Hauptbenutzer für Kunde "{kunde.firmierung}"',
        entity_type='Kunde',
        entity_id=id
    )

    return jsonify({
        'success': True,
        'message': f'{user.full_name} ist jetzt Hauptbenutzer'
    })


# ===== Lead Import & Conversion =====

def _resolve_branche(identifier: str) -> Branche | None:
    """Resolve branche by UUID or name (case-insensitive)."""
    identifier = identifier.strip()
    if not identifier:
        return None

    # Try UUID first (format: xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx)
    if len(identifier) == 36 and identifier.count('-') == 4:
        branche = Branche.query.filter_by(uuid=identifier).first()
        if branche:
            return branche

    # Fallback to name (case-insensitive)
    return Branche.query.filter(
        func.lower(Branche.name) == func.lower(identifier)
    ).first()


def _resolve_verband(kuerzel: str) -> Verband | None:
    """Resolve verband by Kürzel (case-insensitive)."""
    kuerzel = kuerzel.strip()
    if not kuerzel:
        return None

    return Verband.query.filter(
        func.lower(Verband.kuerzel) == func.lower(kuerzel)
    ).first()


def _validate_branchen_consistency(branchen: list[Branche]) -> tuple[bool, str | None]:
    """Check if all branches belong to the same Hauptbranche.

    Returns:
        (is_valid, error_message)
    """
    if not branchen:
        return True, None

    parent_ids = {b.parent_id for b in branchen}

    # All branches must be Unterbranchen (have a parent)
    if None in parent_ids:
        hauptbranchen = [b.name for b in branchen if b.parent_id is None]
        return False, f"Hauptbranche(n) nicht erlaubt: {', '.join(hauptbranchen)}"

    # All must have the same parent
    if len(parent_ids) > 1:
        parents = {Branche.query.get(pid).name for pid in parent_ids}
        return False, f"Inkonsistente Hauptbranchen: {', '.join(parents)}"

    return True, None


@kunden_bp.route('/import', methods=['GET', 'POST'])
@login_required
@mitarbeiter_required
def lead_import():
    """Import leads from CSV file with Branchen and Verbände support.

    CSV format (semicolon-separated, UTF-8):
    firmierung;email;telefon;strasse;plz;ort;website_url;notizen;branchen;verbaende

    - firmierung is required, other fields are optional
    - branchen: Pipe-separated list of branch UUIDs or names (e.g., "Spielwaren|Modellbahn")
    - verbaende: Pipe-separated list of association abbreviations (e.g., "VEDES|EK")
    - First branch is marked as primary (ist_primaer=True)
    """
    from flask import session
    import csv
    import io

    if request.method == 'GET':
        # Load available Branchen and Verbände for documentation
        branchen = Branche.query.filter(
            Branche.aktiv == True,
            Branche.parent_id.isnot(None)  # Only Unterbranchen
        ).order_by(Branche.name).all()
        verbaende = Verband.query.filter_by(aktiv=True).order_by(Verband.name).all()

        return render_template(
            'kunden/import.html',
            branchen=branchen,
            verbaende=verbaende
        )

    # POST: Process CSV upload
    if 'csv_file' not in request.files:
        flash('Keine Datei ausgewählt', 'warning')
        return redirect(url_for('kunden.lead_import'))

    file = request.files['csv_file']
    if file.filename == '':
        flash('Keine Datei ausgewählt', 'warning')
        return redirect(url_for('kunden.lead_import'))

    if not file.filename.endswith('.csv'):
        flash('Nur CSV-Dateien erlaubt', 'danger')
        return redirect(url_for('kunden.lead_import'))

    try:
        content = file.read().decode('utf-8')
        reader = csv.DictReader(io.StringIO(content), delimiter=';')

        # Track import results
        imported = []  # Successfully imported leads
        skipped = []   # Duplicates
        failed = []    # Errors (with details)

        for row_num, row in enumerate(reader, start=2):  # Start at 2 (header is row 1)
            errors = []
            firmierung = (row.get('firmierung') or '').strip()

            # 1. Validate firmierung (required)
            if not firmierung:
                failed.append({
                    'line': row_num,
                    'row': row,
                    'errors': ['Firmierung fehlt']
                })
                continue

            # 2. Check for duplicates
            existing = Kunde.query.filter(
                func.lower(Kunde.firmierung) == func.lower(firmierung)
            ).first()

            if existing:
                skipped.append({
                    'line': row_num,
                    'row': row,
                    'reason': f'Existiert bereits (ID: {existing.id})'
                })
                continue

            # 3. Resolve Branchen
            branchen_resolved = []
            branchen_str = (row.get('branchen') or '').strip()
            if branchen_str:
                for b_str in branchen_str.split('|'):
                    b_str = b_str.strip()
                    if not b_str:
                        continue
                    branche = _resolve_branche(b_str)
                    if not branche:
                        errors.append(f"Branche '{b_str}' nicht gefunden")
                    else:
                        branchen_resolved.append(branche)

            # 4. Validate Branchen consistency (all must have same parent)
            if branchen_resolved:
                is_valid, consistency_error = _validate_branchen_consistency(branchen_resolved)
                if not is_valid:
                    errors.append(consistency_error)

            # 5. Resolve Verbände
            verbaende_resolved = []
            verbaende_str = (row.get('verbaende') or '').strip()
            if verbaende_str:
                for v_str in verbaende_str.split('|'):
                    v_str = v_str.strip()
                    if not v_str:
                        continue
                    verband = _resolve_verband(v_str)
                    if not verband:
                        errors.append(f"Verband '{v_str}' nicht gefunden")
                    else:
                        verbaende_resolved.append(verband)

            # 6. If errors, add to failed list
            if errors:
                failed.append({
                    'line': row_num,
                    'row': row,
                    'errors': errors
                })
                continue

            # 7. Create new Lead
            lead = Kunde(
                firmierung=firmierung,
                email=(row.get('email') or '').strip() or None,
                telefon=(row.get('telefon') or '').strip() or None,
                strasse=(row.get('strasse') or '').strip() or None,
                plz=(row.get('plz') or '').strip() or None,
                ort=(row.get('ort') or '').strip() or None,
                website_url=(row.get('website_url') or '').strip() or None,
                notizen=(row.get('notizen') or '').strip() or None,
                typ=KundeTyp.LEAD.value,
                aktiv=True
            )
            db.session.add(lead)
            db.session.flush()  # Get ID for associations

            # 8. Create Branchen associations
            for i, branche in enumerate(branchen_resolved):
                kb = KundeBranche(
                    kunde_id=lead.id,
                    branche_id=branche.id,
                    ist_primaer=(i == 0)  # First is primary
                )
                db.session.add(kb)

            # 9. Create Verbands associations
            for verband in verbaende_resolved:
                kv = KundeVerband(
                    kunde_id=lead.id,
                    verband_id=verband.id
                )
                db.session.add(kv)

            imported.append({
                'lead': lead,
                'branchen': [b.name for b in branchen_resolved],
                'verbaende': [v.kuerzel for v in verbaende_resolved]
            })

        db.session.commit()

        # Log the import
        log_mittel(
            modul='kunden',
            aktion='leads_importiert',
            details=f'{len(imported)} importiert, {len(skipped)} Duplikate, {len(failed)} Fehler',
            entity_type='Kunde',
            entity_id=None
        )

        # Store failed rows in session for XLSX download
        if failed:
            session['import_failed_rows'] = failed

        # Prepare data for result template
        failed_rows = [
            {
                'line_number': f['line'],
                'row': f['row'],
                'errors': f['errors']
            }
            for f in failed
        ]

        skipped_rows = [
            {
                'line_number': s['line'],
                'row': s['row'],
                'reason': s['reason']
            }
            for s in skipped
        ]

        # Render result page
        return render_template(
            'kunden/import_result.html',
            imported_count=len(imported),
            skipped_count=len(skipped),
            failed_count=len(failed),
            failed_rows=failed_rows,
            skipped_rows=skipped_rows
        )

    except Exception as e:
        db.session.rollback()
        flash(f'Fehler beim Import: {str(e)}', 'danger')
        return redirect(url_for('kunden.lead_import'))


@kunden_bp.route('/import/fehler-xlsx')
@login_required
@mitarbeiter_required
def lead_import_fehler_xlsx():
    """Download failed import rows as XLSX."""
    from flask import session, Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io

    failed_rows = session.get('import_failed_rows', [])
    if not failed_rows:
        flash('Keine fehlerhaften Zeilen vorhanden', 'info')
        return redirect(url_for('kunden.lead_import'))

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Import-Fehler"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")

    # Headers
    headers = ['Zeile', 'Firmierung', 'Email', 'Branchen', 'Verbände', 'Fehler']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for i, item in enumerate(failed_rows, start=2):
        row = item.get('row', {})
        ws.cell(row=i, column=1, value=item.get('line', ''))
        ws.cell(row=i, column=2, value=row.get('firmierung', ''))
        ws.cell(row=i, column=3, value=row.get('email', ''))
        ws.cell(row=i, column=4, value=row.get('branchen', ''))
        ws.cell(row=i, column=5, value=row.get('verbaende', ''))
        ws.cell(row=i, column=6, value='; '.join(item.get('errors', [])))

    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 50

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Clear session data
    session.pop('import_failed_rows', None)

    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=import_fehler.xlsx'}
    )


@kunden_bp.route('/<int:id>/konvertieren', methods=['POST'])
@login_required
@mitarbeiter_required
def konvertieren(id):
    """Convert a Lead to a Kunde.

    Simply changes the typ field from 'lead' to 'kunde'.
    """
    kunde = Kunde.query.get_or_404(id)

    if kunde.typ != KundeTyp.LEAD.value:
        flash('Nur Leads können zu Kunden konvertiert werden', 'warning')
        return redirect(url_for('kunden.detail', id=id))

    kunde.typ = KundeTyp.KUNDE.value
    db.session.commit()

    # Log the conversion
    log_mittel(
        modul='kunden',
        aktion='lead_konvertiert',
        details=f'Lead "{kunde.firmierung}" wurde zu Kunde konvertiert',
        entity_type='Kunde',
        entity_id=id
    )

    flash(f'"{kunde.firmierung}" wurde zu Kunde konvertiert', 'success')
    return redirect(url_for('kunden.detail', id=id))
