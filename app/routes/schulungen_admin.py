"""Admin routes for Schulungen management (PRD-010).

Blueprint: schulungen_admin_bp
Prefix: /admin/schulungen/

This module provides routes for administrators to manage:
- Schulungen (training templates)
- Schulungsthemen (reusable topics)
- Schulungsdurchführungen (concrete executions)
- Schulungsbuchungen (bookings)
- Excel export for ERP integration
"""
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from flask import Blueprint, render_template, request, redirect, url_for, flash, abort, send_file
from flask_login import login_required, current_user

from app import db
from app.models import (
    Schulung, Schulungsthema, SchulungThema,
    Schulungsdurchfuehrung, Schulungstermin,
    Schulungsbuchung, BuchungStatus, DurchfuehrungStatus
)
from app.services import log_event, log_hoch
from app.routes.admin import mitarbeiter_required

schulungen_admin_bp = Blueprint('schulungen_admin', __name__, url_prefix='/admin/schulungen')


# =============================================================================
# SCHULUNGEN (Training Templates)
# =============================================================================

@schulungen_admin_bp.route('/')
@login_required
@mitarbeiter_required
def index():
    """Dashboard with overview of all trainings."""
    schulungen = Schulung.query.order_by(Schulung.sortierung, Schulung.titel).all()
    themen = Schulungsthema.query.filter_by(aktiv=True).order_by(Schulungsthema.titel).all()

    # Stats
    stats = {
        'schulungen_aktiv': sum(1 for s in schulungen if s.aktiv),
        'schulungen_inaktiv': sum(1 for s in schulungen if not s.aktiv),
        'themen_count': len(themen),
        'buchungen_offen': Schulungsbuchung.query.filter_by(
            status=BuchungStatus.GEBUCHT.value
        ).count(),
        'warteliste_count': Schulungsbuchung.query.filter_by(
            status=BuchungStatus.WARTELISTE.value
        ).count(),
    }

    return render_template(
        'administration/schulungen/index.html',
        schulungen=schulungen,
        themen=themen,
        stats=stats,
        admin_tab='module'
    )


@schulungen_admin_bp.route('/neu', methods=['GET', 'POST'])
@login_required
@mitarbeiter_required
def schulung_neu():
    """Create a new training."""
    themen = Schulungsthema.query.filter_by(aktiv=True).order_by(Schulungsthema.titel).all()

    if request.method == 'POST':
        titel = request.form.get('titel', '').strip()
        beschreibung = request.form.get('beschreibung', '').strip()
        artikelnummer = request.form.get('artikelnummer', '').strip()
        preis = request.form.get('preis', '0').replace(',', '.')
        sonderpreis = request.form.get('sonderpreis', '').replace(',', '.').strip()
        aktionszeitraum_von = request.form.get('aktionszeitraum_von', '').strip()
        aktionszeitraum_bis = request.form.get('aktionszeitraum_bis', '').strip()
        max_teilnehmer = request.form.get('max_teilnehmer', '10')
        storno_frist_tage = request.form.get('storno_frist_tage', '7')
        aktiv = request.form.get('aktiv') == 'on'
        thema_ids = request.form.getlist('thema_ids')

        # Validation
        errors = []
        if not titel:
            errors.append('Bitte geben Sie einen Titel ein.')
        try:
            preis_decimal = Decimal(preis)
            if preis_decimal <= 0:
                errors.append('Der Preis muss größer als 0 sein.')
        except Exception:
            errors.append('Bitte geben Sie einen gültigen Preis ein.')

        if errors:
            for error in errors:
                flash(error, 'danger')
            return render_template(
                'administration/schulungen/schulung_form.html',
                schulung=None,
                themen=themen,
                form_data=request.form,
                admin_tab='module'
            )

        # Create Schulung
        schulung = Schulung(
            titel=titel,
            beschreibung=beschreibung or None,
            artikelnummer=artikelnummer or None,
            preis=Decimal(preis),
            sonderpreis=Decimal(sonderpreis) if sonderpreis else None,
            aktionszeitraum_von=datetime.strptime(aktionszeitraum_von, '%Y-%m-%d').date() if aktionszeitraum_von else None,
            aktionszeitraum_bis=datetime.strptime(aktionszeitraum_bis, '%Y-%m-%d').date() if aktionszeitraum_bis else None,
            max_teilnehmer=int(max_teilnehmer),
            storno_frist_tage=int(storno_frist_tage),
            aktiv=aktiv
        )
        db.session.add(schulung)
        db.session.flush()

        # Add Themen
        for idx, thema_id in enumerate(thema_ids):
            if thema_id:
                verknuepfung = SchulungThema(
                    schulung_id=schulung.id,
                    thema_id=int(thema_id),
                    sortierung=idx
                )
                db.session.add(verknuepfung)

        db.session.commit()

        log_hoch(
            'schulungen', 'erstellt',
            f'Schulung "{titel}" erstellt',
            entity_type='Schulung', entity_id=schulung.id
        )
        flash(f'Schulung "{titel}" wurde erfolgreich erstellt.', 'success')
        return redirect(url_for('schulungen_admin.index'))

    return render_template(
        'administration/schulungen/schulung_form.html',
        schulung=None,
        themen=themen,
        form_data={},
        admin_tab='module'
    )


@schulungen_admin_bp.route('/<int:id>', methods=['GET', 'POST'])
@login_required
@mitarbeiter_required
def schulung_bearbeiten(id):
    """Edit an existing training."""
    schulung = Schulung.query.get_or_404(id)
    themen = Schulungsthema.query.filter_by(aktiv=True).order_by(Schulungsthema.titel).all()

    if request.method == 'POST':
        schulung.titel = request.form.get('titel', '').strip()
        schulung.beschreibung = request.form.get('beschreibung', '').strip() or None
        schulung.artikelnummer = request.form.get('artikelnummer', '').strip() or None

        preis = request.form.get('preis', '0').replace(',', '.')
        sonderpreis = request.form.get('sonderpreis', '').replace(',', '.').strip()
        aktionszeitraum_von = request.form.get('aktionszeitraum_von', '').strip()
        aktionszeitraum_bis = request.form.get('aktionszeitraum_bis', '').strip()

        schulung.preis = Decimal(preis)
        schulung.sonderpreis = Decimal(sonderpreis) if sonderpreis else None
        schulung.aktionszeitraum_von = datetime.strptime(aktionszeitraum_von, '%Y-%m-%d').date() if aktionszeitraum_von else None
        schulung.aktionszeitraum_bis = datetime.strptime(aktionszeitraum_bis, '%Y-%m-%d').date() if aktionszeitraum_bis else None

        schulung.max_teilnehmer = int(request.form.get('max_teilnehmer', '10'))
        schulung.storno_frist_tage = int(request.form.get('storno_frist_tage', '7'))
        schulung.aktiv = request.form.get('aktiv') == 'on'

        # Update Themen
        thema_ids = request.form.getlist('thema_ids')

        # Remove old links
        SchulungThema.query.filter_by(schulung_id=schulung.id).delete()

        # Add new links
        for idx, thema_id in enumerate(thema_ids):
            if thema_id:
                verknuepfung = SchulungThema(
                    schulung_id=schulung.id,
                    thema_id=int(thema_id),
                    sortierung=idx
                )
                db.session.add(verknuepfung)

        db.session.commit()

        log_event(
            'schulungen', 'bearbeitet',
            f'Schulung "{schulung.titel}" bearbeitet',
            entity_type='Schulung', entity_id=schulung.id
        )
        flash(f'Schulung "{schulung.titel}" wurde aktualisiert.', 'success')
        return redirect(url_for('schulungen_admin.index'))

    return render_template(
        'administration/schulungen/schulung_form.html',
        schulung=schulung,
        themen=themen,
        form_data={},
        admin_tab='module'
    )


@schulungen_admin_bp.route('/<int:id>/loeschen', methods=['POST'])
@login_required
@mitarbeiter_required
def schulung_loeschen(id):
    """Delete a training (if no bookings exist)."""
    schulung = Schulung.query.get_or_404(id)

    # Check for existing bookings
    buchungen_count = Schulungsbuchung.query.join(Schulungsdurchfuehrung).filter(
        Schulungsdurchfuehrung.schulung_id == schulung.id
    ).count()

    if buchungen_count > 0:
        flash(f'Schulung kann nicht gelöscht werden: Es existieren {buchungen_count} Buchungen.', 'danger')
        return redirect(url_for('schulungen_admin.index'))

    titel = schulung.titel
    db.session.delete(schulung)
    db.session.commit()

    log_hoch(
        'schulungen', 'geloescht',
        f'Schulung "{titel}" gelöscht',
        entity_type='Schulung', entity_id=id
    )
    flash(f'Schulung "{titel}" wurde gelöscht.', 'success')
    return redirect(url_for('schulungen_admin.index'))


# =============================================================================
# SCHULUNGSTHEMEN (Reusable Topics)
# =============================================================================

@schulungen_admin_bp.route('/themen')
@login_required
@mitarbeiter_required
def themen_liste():
    """List all training topics."""
    themen = Schulungsthema.query.order_by(Schulungsthema.titel).all()
    return render_template(
        'administration/schulungen/themen_liste.html',
        themen=themen,
        admin_tab='module'
    )


@schulungen_admin_bp.route('/themen/neu', methods=['GET', 'POST'])
@login_required
@mitarbeiter_required
def thema_neu():
    """Create a new topic."""
    if request.method == 'POST':
        titel = request.form.get('titel', '').strip()
        beschreibung = request.form.get('beschreibung', '').strip()
        dauer_minuten = request.form.get('dauer_minuten', '45')
        aktiv = request.form.get('aktiv') == 'on'

        if not titel:
            flash('Bitte geben Sie einen Titel ein.', 'danger')
            return render_template(
                'administration/schulungen/thema_form.html',
                thema=None,
                form_data=request.form,
                admin_tab='module'
            )

        thema = Schulungsthema(
            titel=titel,
            beschreibung=beschreibung or None,
            dauer_minuten=int(dauer_minuten),
            aktiv=aktiv
        )
        db.session.add(thema)
        db.session.commit()

        log_event(
            'schulungen', 'erstellt',
            f'Thema "{titel}" erstellt',
            entity_type='Schulungsthema', entity_id=thema.id
        )
        flash(f'Thema "{titel}" wurde erstellt.', 'success')
        return redirect(url_for('schulungen_admin.themen_liste'))

    return render_template(
        'administration/schulungen/thema_form.html',
        thema=None,
        form_data={},
        admin_tab='module'
    )


@schulungen_admin_bp.route('/themen/<int:id>', methods=['GET', 'POST'])
@login_required
@mitarbeiter_required
def thema_bearbeiten(id):
    """Edit a topic."""
    thema = Schulungsthema.query.get_or_404(id)

    if request.method == 'POST':
        thema.titel = request.form.get('titel', '').strip()
        thema.beschreibung = request.form.get('beschreibung', '').strip() or None
        thema.dauer_minuten = int(request.form.get('dauer_minuten', '45'))
        thema.aktiv = request.form.get('aktiv') == 'on'

        db.session.commit()

        log_event(
            'schulungen', 'bearbeitet',
            f'Thema "{thema.titel}" bearbeitet',
            entity_type='Schulungsthema', entity_id=thema.id
        )
        flash(f'Thema "{thema.titel}" wurde aktualisiert.', 'success')
        return redirect(url_for('schulungen_admin.themen_liste'))

    return render_template(
        'administration/schulungen/thema_form.html',
        thema=thema,
        form_data={},
        admin_tab='module'
    )


@schulungen_admin_bp.route('/themen/<int:id>/duplizieren', methods=['POST'])
@login_required
@mitarbeiter_required
def thema_duplizieren(id):
    """Duplicate a topic."""
    original = Schulungsthema.query.get_or_404(id)

    kopie = Schulungsthema(
        titel=f"{original.titel} (Kopie)",
        beschreibung=original.beschreibung,
        dauer_minuten=original.dauer_minuten,
        aktiv=False  # Kopie standardmäßig inaktiv
    )
    db.session.add(kopie)
    db.session.commit()

    log_event(
        'schulungen', 'dupliziert',
        f'Thema "{original.titel}" dupliziert zu "{kopie.titel}"',
        entity_type='Schulungsthema', entity_id=kopie.id
    )
    flash(f'Thema "{original.titel}" wurde dupliziert.', 'success')
    return redirect(url_for('schulungen_admin.thema_bearbeiten', id=kopie.id))


# =============================================================================
# DURCHFÜHRUNGEN (Concrete Executions)
# =============================================================================

@schulungen_admin_bp.route('/durchfuehrungen')
@login_required
@mitarbeiter_required
def durchfuehrungen_liste():
    """List all executions with filters."""
    status_filter = request.args.get('status', '')
    schulung_filter = request.args.get('schulung', '')

    query = Schulungsdurchfuehrung.query

    if status_filter:
        query = query.filter_by(status=status_filter)
    if schulung_filter:
        query = query.filter_by(schulung_id=int(schulung_filter))

    durchfuehrungen = query.order_by(Schulungsdurchfuehrung.start_datum.desc()).all()
    schulungen = Schulung.query.filter_by(aktiv=True).order_by(Schulung.titel).all()

    return render_template(
        'administration/schulungen/durchfuehrungen_liste.html',
        durchfuehrungen=durchfuehrungen,
        schulungen=schulungen,
        status_filter=status_filter,
        schulung_filter=schulung_filter,
        DurchfuehrungStatus=DurchfuehrungStatus,
        admin_tab='module'
    )


@schulungen_admin_bp.route('/durchfuehrungen/neu', methods=['GET', 'POST'])
@schulungen_admin_bp.route('/durchfuehrungen/neu/<int:schulung_id>', methods=['GET', 'POST'])
@login_required
@mitarbeiter_required
def durchfuehrung_neu(schulung_id=None):
    """Create a new execution."""
    schulungen = Schulung.query.filter_by(aktiv=True).order_by(Schulung.titel).all()

    if request.method == 'POST':
        schulung_id = int(request.form.get('schulung_id'))
        start_datum = request.form.get('start_datum', '').strip()
        teams_link = request.form.get('teams_link', '').strip()
        anmerkungen = request.form.get('anmerkungen', '').strip()

        # Parse Terminmuster
        wochentage = request.form.getlist('wochentage')
        uhrzeit = request.form.get('uhrzeit', '14:00').strip()

        if not start_datum:
            flash('Bitte geben Sie ein Startdatum ein.', 'danger')
            return render_template(
                'administration/schulungen/durchfuehrung_form.html',
                durchfuehrung=None,
                schulungen=schulungen,
                vorausgewaehlte_schulung_id=schulung_id,
                form_data=request.form,
                admin_tab='module'
            )

        terminmuster = {
            'wochentage': wochentage,
            'uhrzeit': uhrzeit
        }

        durchfuehrung = Schulungsdurchfuehrung(
            schulung_id=schulung_id,
            start_datum=datetime.strptime(start_datum, '%Y-%m-%d').date(),
            terminmuster=terminmuster,
            teams_link=teams_link or None,
            anmerkungen=anmerkungen or None,
            status=DurchfuehrungStatus.GEPLANT.value
        )
        db.session.add(durchfuehrung)
        db.session.commit()

        log_event(
            'schulungen', 'erstellt',
            f'Durchführung für "{durchfuehrung.schulung.titel}" ab {start_datum} erstellt',
            entity_type='Schulungsdurchfuehrung', entity_id=durchfuehrung.id
        )
        flash('Durchführung wurde erstellt.', 'success')
        return redirect(url_for('schulungen_admin.durchfuehrung_detail', id=durchfuehrung.id))

    return render_template(
        'administration/schulungen/durchfuehrung_form.html',
        durchfuehrung=None,
        schulungen=schulungen,
        vorausgewaehlte_schulung_id=schulung_id,
        form_data={},
        admin_tab='module'
    )


@schulungen_admin_bp.route('/durchfuehrungen/<int:id>')
@login_required
@mitarbeiter_required
def durchfuehrung_detail(id):
    """View execution details with participants."""
    durchfuehrung = Schulungsdurchfuehrung.query.get_or_404(id)

    return render_template(
        'administration/schulungen/durchfuehrung_detail.html',
        durchfuehrung=durchfuehrung,
        BuchungStatus=BuchungStatus,
        DurchfuehrungStatus=DurchfuehrungStatus,
        admin_tab='module'
    )


@schulungen_admin_bp.route('/durchfuehrungen/<int:id>/status', methods=['POST'])
@login_required
@mitarbeiter_required
def durchfuehrung_status(id):
    """Change execution status."""
    durchfuehrung = Schulungsdurchfuehrung.query.get_or_404(id)
    neuer_status = request.form.get('status')

    try:
        if neuer_status == DurchfuehrungStatus.AKTIV.value:
            durchfuehrung.aktivieren()
        elif neuer_status == DurchfuehrungStatus.ABGESCHLOSSEN.value:
            durchfuehrung.abschliessen()
        elif neuer_status == DurchfuehrungStatus.ABGESAGT.value:
            durchfuehrung.absagen()

        db.session.commit()
        log_event(
            'schulungen', 'status_geaendert',
            f'Status auf "{neuer_status}" geändert',
            entity_type='Schulungsdurchfuehrung', entity_id=durchfuehrung.id
        )
        flash(f'Status wurde auf "{neuer_status}" geändert.', 'success')
    except ValueError as e:
        flash(str(e), 'danger')

    return redirect(url_for('schulungen_admin.durchfuehrung_detail', id=id))


# =============================================================================
# BUCHUNGEN (Bookings)
# =============================================================================

@schulungen_admin_bp.route('/buchungen')
@login_required
@mitarbeiter_required
def buchungen_liste():
    """List all bookings with filters."""
    status_filter = request.args.get('status', '')
    durchfuehrung_filter = request.args.get('durchfuehrung', '')

    query = Schulungsbuchung.query

    if status_filter:
        query = query.filter_by(status=status_filter)
    if durchfuehrung_filter:
        query = query.filter_by(durchfuehrung_id=int(durchfuehrung_filter))

    buchungen = query.order_by(Schulungsbuchung.gebucht_am.desc()).all()
    durchfuehrungen = Schulungsdurchfuehrung.get_kommende()

    return render_template(
        'administration/schulungen/buchungen_liste.html',
        buchungen=buchungen,
        durchfuehrungen=durchfuehrungen,
        status_filter=status_filter,
        durchfuehrung_filter=durchfuehrung_filter,
        BuchungStatus=BuchungStatus,
        admin_tab='module'
    )


@schulungen_admin_bp.route('/buchungen/<int:id>/freischalten', methods=['POST'])
@login_required
@mitarbeiter_required
def buchung_freischalten(id):
    """Move booking from waitlist to booked."""
    buchung = Schulungsbuchung.query.get_or_404(id)

    try:
        buchung.von_warteliste_freischalten()
        db.session.commit()

        log_hoch(
            'schulungen', 'freigeschaltet',
            f'Buchung von Warteliste freigeschaltet für Kunde {buchung.kunde.firmierung}',
            entity_type='Schulungsbuchung', entity_id=buchung.id
        )
        flash(f'Buchung für {buchung.kunde.firmierung} wurde freigeschaltet.', 'success')
    except ValueError as e:
        flash(str(e), 'danger')

    return redirect(url_for('schulungen_admin.durchfuehrung_detail', id=buchung.durchfuehrung_id))


@schulungen_admin_bp.route('/buchungen/<int:id>/stornieren', methods=['POST'])
@login_required
@mitarbeiter_required
def buchung_stornieren(id):
    """Cancel a booking (admin action)."""
    buchung = Schulungsbuchung.query.get_or_404(id)

    try:
        buchung.stornieren()
        db.session.commit()

        log_hoch(
            'schulungen', 'admin_storniert',
            f'Buchung durch Admin storniert für Kunde {buchung.kunde.firmierung}',
            entity_type='Schulungsbuchung', entity_id=buchung.id
        )
        flash(f'Buchung für {buchung.kunde.firmierung} wurde storniert.', 'success')
    except ValueError as e:
        flash(str(e), 'danger')

    return redirect(url_for('schulungen_admin.durchfuehrung_detail', id=buchung.durchfuehrung_id))


# =============================================================================
# EXCEL EXPORT (ERP Integration)
# =============================================================================

@schulungen_admin_bp.route('/export')
@login_required
@mitarbeiter_required
def export_buchungen():
    """Export bookings to Excel for ERP integration."""
    von_datum = request.args.get('von', '')
    bis_datum = request.args.get('bis', '')

    von = datetime.strptime(von_datum, '%Y-%m-%d').date() if von_datum else None
    bis = datetime.strptime(bis_datum, '%Y-%m-%d').date() if bis_datum else None

    buchungen = Schulungsbuchung.get_fuer_export(von_datum=von, bis_datum=bis)

    # If no filter provided, show the export form
    if not von_datum and not bis_datum:
        return render_template(
            'administration/schulungen/export.html',
            buchungen=[],
            von_datum='',
            bis_datum='',
            admin_tab='module'
        )

    # Generate Excel
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Schulungsbuchungen'

        # Header
        headers = ['Kundennummer', 'Firmenname', 'Schulung', 'Artikelnummer', 'Preis', 'Buchungsdatum', 'Start-Datum']
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Data rows
        for row, buchung in enumerate(buchungen, 2):
            ws.cell(row=row, column=1, value=buchung.kunde.kundennummer or '')
            ws.cell(row=row, column=2, value=buchung.kunde.firmierung)
            ws.cell(row=row, column=3, value=buchung.schulung.titel)
            ws.cell(row=row, column=4, value=buchung.schulung.artikelnummer or '')
            ws.cell(row=row, column=5, value=float(buchung.preis_bei_buchung))
            ws.cell(row=row, column=6, value=buchung.gebucht_am.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=7, value=buchung.durchfuehrung.start_datum.strftime('%Y-%m-%d'))

        # Adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'schulungsbuchungen_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        log_event(
            'schulungen', 'export',
            f'Excel-Export erstellt ({len(buchungen)} Buchungen)'
        )

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except ImportError:
        flash('openpyxl ist nicht installiert. Bitte installieren Sie es mit: uv add openpyxl', 'danger')
        return redirect(url_for('schulungen_admin.buchungen_liste'))


# =============================================================================
# EINSTELLUNGEN (Module Settings)
# =============================================================================

@schulungen_admin_bp.route('/einstellungen', methods=['GET', 'POST'])
@login_required
@mitarbeiter_required
def einstellungen():
    """Schulungen module settings page."""
    from app.models import Config

    if request.method == 'POST':
        # Save settings
        for key in request.form:
            if key.startswith('schulungen_'):
                Config.set_value(key, request.form[key])

        flash('Einstellungen gespeichert', 'success')
        return redirect(url_for('schulungen_admin.einstellungen'))

    # Load current settings
    settings = {
        'max_teilnehmer_default': Config.get_value('schulungen_max_teilnehmer_default', '10'),
        'storno_frist_tage_default': Config.get_value('schulungen_storno_frist_tage_default', '7'),
        'buchungsbestaetigung_email': Config.get_value('schulungen_buchungsbestaetigung_email', 'true'),
    }

    return render_template(
        'administration/schulungen/einstellungen.html',
        settings=settings,
        admin_tab='module'
    )
